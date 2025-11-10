import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Manual Test Plan"

# Define header styling
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Define cell styling
module_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
module_font = Font(bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 60

# Add headers
headers = ["Feature/Module", "Action to Test", "Expected Result"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Test cases data
test_cases = [
    # Authentication Module
    ("Authentication", "User Registration - Valid Data", "User is successfully registered with email and password. User is redirected to login page. Password is hashed in database."),
    ("Authentication", "User Registration - Duplicate Email", "Registration fails with error message 'Email already exists'. User remains on registration page."),
    ("Authentication", "User Registration - Invalid Email Format", "Registration fails with validation error. Email field shows error message."),
    ("Authentication", "User Registration - Password Less Than 6 Characters", "Registration fails with error 'Password must be at least 6 characters'. Form shows validation error."),
    ("Authentication", "User Login - Valid Credentials", "User is successfully logged in. JWT token is generated. User is redirected to dashboard. Session is maintained."),
    ("Authentication", "User Login - Invalid Email", "Login fails with error 'Invalid credentials'. User remains on login page."),
    ("Authentication", "User Login - Invalid Password", "Login fails with error 'Invalid credentials'. User remains on login page."),
    ("Authentication", "Access Protected Route Without Login", "User is redirected to login page. Protected content is not accessible."),
    ("Authentication", "Token Expiration", "After token expires, user is logged out and redirected to login page."),
    ("Authentication", "Logout Functionality", "User session is cleared. User is redirected to login page. Subsequent API calls require re-authentication."),

    # Lead Management
    ("Lead Management", "Create New Lead - All Fields", "Lead is created with name, phone, alternate phone, source, referrer. Lead appears in leads list. Success message is shown."),
    ("Lead Management", "Create New Lead - Required Fields Only", "Lead is created with only name and phone. Lead appears in leads list with default source 'unknown'."),
    ("Lead Management", "Create Lead - Duplicate Phone Number", "Lead is created successfully (system allows duplicate leads for different users or tracking)."),
    ("Lead Management", "View All Leads - List Page", "All leads belonging to logged-in user are displayed. Leads are paginated. Columns show name, phone, source, created date."),
    ("Lead Management", "Search Lead by Name", "Search input filters leads by name. Results update dynamically. Only matching leads are displayed."),
    ("Lead Management", "Search Lead by Phone Number", "Search input filters leads by phone number. Partial matches work. Results update in real-time."),
    ("Lead Management", "Filter Leads by Source - IndiaMart", "Only leads with source 'IndiaMart' are displayed. Other sources are hidden. Count matches filtered results."),
    ("Lead Management", "Filter Leads by Source - JustDial", "Only leads with source 'JustDial' are displayed. Filter works correctly."),
    ("Lead Management", "Filter Leads by Source - Referral", "Only leads with source 'Referral' are displayed. Referrer information is shown."),
    ("Lead Management", "Filter Leads by Date Range", "Leads created within selected date range are displayed. Dates outside range are excluded."),
    ("Lead Management", "Edit Existing Lead", "Lead details are updated successfully. Changes reflect immediately in list. Updated timestamp is modified."),
    ("Lead Management", "Delete Lead - No Associated Enquiries", "Lead is deleted successfully. Lead is removed from list. Confirmation prompt appears before deletion."),
    ("Lead Management", "Delete Lead - With Associated Enquiries", "System warns about associated enquiries. Option to delete enquiries or cancel deletion is provided."),
    ("Lead Management", "View Single Lead Details", "All lead information is displayed including name, phone, alternate phone, source, referrer, created date, updated date."),
    ("Lead Management", "Pagination - Navigate Pages", "Pagination controls work correctly. Next/Previous buttons navigate between pages. Page numbers are accurate."),

    # Customer Enquiry Management
    ("Customer Enquiry Management", "Create Enquiry from Lead", "New enquiry is created linked to selected lead. Enquiry appears in enquiries list with status 'pending'. Lead ID is correctly associated."),
    ("Customer Enquiry Management", "Create Enquiry - All Fields", "Enquiry is created with from location, to location, cargo type, cargo weight, remarks, source, referrer. All data is saved correctly."),
    ("Customer Enquiry Management", "View All Enquiries", "All enquiries for logged-in user are displayed. Pagination works. Columns show lead name, from/to, cargo type, weight, status."),
    ("Customer Enquiry Management", "Search Enquiries by Customer Name", "Search filters enquiries by associated lead name. Results update dynamically."),
    ("Customer Enquiry Management", "Filter by Enquiry Status - Pending", "Only enquiries with status 'pending' are displayed. Count is accurate."),
    ("Customer Enquiry Management", "Filter by Enquiry Status - Accepted", "Only accepted enquiries are shown. Filter works correctly."),
    ("Customer Enquiry Management", "Filter by Enquiry Status - Rejected", "Only rejected enquiries are displayed."),
    ("Customer Enquiry Management", "Filter by Source - IndiaMart", "Enquiries from IndiaMart source are filtered correctly."),
    ("Customer Enquiry Management", "Filter by Location (From)", "Enquiries with matching 'from' location are displayed. Partial text matching works."),
    ("Customer Enquiry Management", "Filter by Location (To)", "Enquiries with matching 'to' location are displayed."),
    ("Customer Enquiry Management", "Filter by Cargo Type", "Enquiries with specific cargo type are filtered. Results match exactly."),
    ("Customer Enquiry Management", "Filter by Weight Range - Min Weight", "Enquiries with weight >= minimum value are displayed. Lower weights are excluded."),
    ("Customer Enquiry Management", "Filter by Weight Range - Max Weight", "Enquiries with weight <= maximum value are displayed. Higher weights are excluded."),
    ("Customer Enquiry Management", "Filter by Date Range", "Enquiries created within date range are shown. Correct filtering by creation date."),
    ("Customer Enquiry Management", "Filter - Has Customer Order (Yes)", "Only enquiries with associated customer orders are displayed. Enquiries without orders are hidden."),
    ("Customer Enquiry Management", "Filter - Has Customer Order (No)", "Only enquiries without customer orders are displayed."),
    ("Customer Enquiry Management", "Expand Enquiry Row - View Quotes", "Expandable row shows all quotes associated with enquiry. Quote amounts, margins, status displayed."),
    ("Customer Enquiry Management", "Expand Enquiry Row - View Transport Orders", "Expandable row shows all transport orders for enquiry. Broker name, route, amount displayed."),
    ("Customer Enquiry Management", "Edit Enquiry Details", "Enquiry is updated successfully. Changes reflect in list. Updated timestamp is modified."),
    ("Customer Enquiry Management", "Change Enquiry Status - Accept", "Enquiry status changes to 'accepted'. Status badge updates. Timestamp records change."),
    ("Customer Enquiry Management", "Change Enquiry Status - Reject", "Enquiry status changes to 'rejected'. Status updates correctly."),
    ("Customer Enquiry Management", "Delete Enquiry - No Dependencies", "Enquiry is deleted successfully. Removed from list. Confirmation prompt shown."),
    ("Customer Enquiry Management", "Delete Enquiry - With Quotes", "System warns about associated quotes. Options to proceed or cancel are provided."),
    ("Customer Enquiry Management", "Combine Multiple Filters", "Multiple filters work together (e.g., status + location + date range). Results match all criteria."),

    # Broker Management
    ("Broker Management", "Create New Broker - All Details", "Broker is created with company name, person name, phone, alternate phone, city, remarks. Broker appears in list."),
    ("Broker Management", "Create Broker - Required Fields Only", "Broker is created with minimum required fields (company name, person name, phone). Record is saved."),
    ("Broker Management", "View All Brokers List", "All brokers for user are displayed. Pagination works. Columns show company name, person name, phone, city."),
    ("Broker Management", "Search Broker by Company Name", "Search filters brokers by company name. Results update dynamically."),
    ("Broker Management", "Search Broker by Person Name", "Search filters by contact person name. Partial matches work."),
    ("Broker Management", "Search Broker by Phone Number", "Search filters by phone number. Results are accurate."),
    ("Broker Management", "Search Broker by City", "Search filters brokers by city location. Results update correctly."),
    ("Broker Management", "View Single Broker Details", "All broker information is displayed including regions and vehicle types. Contact details shown correctly."),
    ("Broker Management", "Edit Broker Details", "Broker information is updated successfully. Changes reflect in list immediately."),
    ("Broker Management", "Delete Broker - No Associated Bids", "Broker is deleted successfully. Removed from list. Confirmation shown."),
    ("Broker Management", "Delete Broker - With Associated Bids", "System warns about existing bids. Options to proceed or cancel provided."),
    ("Broker Management", "Add Operating Region to Broker", "New region is added to broker's operating regions. Region appears in broker's region list."),
    ("Broker Management", "Add Multiple Regions to Broker", "Multiple regions can be added. All regions are saved and displayed correctly."),
    ("Broker Management", "Edit Broker Region", "Region name/details are updated. Changes saved successfully."),
    ("Broker Management", "Remove Region from Broker", "Region is removed from broker. Region no longer appears in list. Confirmation prompt shown."),
    ("Broker Management", "Add Vehicle Type to Broker", "New vehicle type is added to broker's available types. Vehicle type appears in list."),
    ("Broker Management", "Add Multiple Vehicle Types", "Multiple vehicle types can be added to broker. All types saved correctly."),
    ("Broker Management", "Edit Vehicle Type for Broker", "Vehicle type is updated. Changes reflect in broker's vehicle list."),
    ("Broker Management", "Remove Vehicle Type from Broker", "Vehicle type is removed from broker. No longer appears in list. Confirmation shown."),
    ("Broker Management", "Pagination in Brokers List", "Pagination controls work. Navigate between pages correctly. Page count accurate."),

    # Transport Broker Rate Enquiries
    ("Transport Broker Rate Enquiries", "Create New Rate Enquiry", "Rate enquiry is created with route, cargo type, cargo weight, transport date, remarks. Status defaults to 'open'. Appears in list."),
    ("Transport Broker Rate Enquiries", "View All Rate Enquiries", "All rate enquiries displayed with route, cargo type, weight, status, L1 & L2 rates. Pagination works."),
    ("Transport Broker Rate Enquiries", "Filter by Status - Open", "Only rate enquiries with status 'open' are displayed."),
    ("Transport Broker Rate Enquiries", "Filter by Status - Bidding", "Only enquiries in 'bidding' status are shown."),
    ("Transport Broker Rate Enquiries", "Filter by Status - Quoted", "Only enquiries with status 'quoted' are displayed."),
    ("Transport Broker Rate Enquiries", "Filter by Status - Closed", "Only closed enquiries are shown."),
    ("Transport Broker Rate Enquiries", "Search by Cargo Type", "Rate enquiries filtered by cargo type. Partial matching works."),
    ("Transport Broker Rate Enquiries", "Search by Route", "Enquiries filtered by route name or locations. Results accurate."),
    ("Transport Broker Rate Enquiries", "View L1 Rate (Lowest Bid)", "L1 rate shows the lowest bid received from brokers. Calculation is accurate. Broker name displayed."),
    ("Transport Broker Rate Enquiries", "View L2 Rate (Second Lowest Bid)", "L2 rate shows second lowest bid. Calculation correct. Broker name shown."),
    ("Transport Broker Rate Enquiries", "Rate Enquiry with No Bids", "L1 and L2 columns show 'No bids yet' or empty. No error occurs."),
    ("Transport Broker Rate Enquiries", "Rate Enquiry with Only One Bid", "L1 shows the single bid. L2 shows 'N/A' or empty. No errors."),
    ("Transport Broker Rate Enquiries", "Edit Rate Enquiry", "Rate enquiry details updated successfully. Changes reflect in list."),
    ("Transport Broker Rate Enquiries", "Change Status - Open to Bidding", "Status updates to 'bidding'. Status badge changes. Timestamp updated."),
    ("Transport Broker Rate Enquiries", "Change Status - Bidding to Quoted", "Status changes to 'quoted'. Indicates bids received and evaluated."),
    ("Transport Broker Rate Enquiries", "Change Status - Quoted to Closed", "Status becomes 'closed'. Enquiry marked as completed."),
    ("Transport Broker Rate Enquiries", "Delete Rate Enquiry - No Bids", "Enquiry deleted successfully. Removed from list. Confirmation shown."),
    ("Transport Broker Rate Enquiries", "Delete Rate Enquiry - With Bids", "Warning about associated bids. Options to delete or cancel provided."),
    ("Transport Broker Rate Enquiries", "View Rate Enquiry Detail Page", "Detail page shows all enquiry info + list of all bids received. Bid details include broker, rate, date."),

    # Broker Bids
    ("Broker Bids", "Submit Bid on Rate Enquiry", "Bid is submitted with broker selection and rate amount. Bid appears in enquiry's bid list. L1/L2 calculations update."),
    ("Broker Bids", "Submit Multiple Bids from Different Brokers", "Multiple brokers can bid on same enquiry. All bids recorded. L1/L2 ranks correctly."),
    ("Broker Bids", "Submit Lower Bid - Becomes L1", "New lowest bid becomes L1 rate. Previous L1 becomes L2. Rankings update automatically."),
    ("Broker Bids", "View All Bids for Rate Enquiry", "All bids displayed with broker name, rate, submission date. Sorted by rate (lowest first)."),
    ("Broker Bids", "Edit Existing Bid", "Bid rate is updated. L1/L2 calculations refresh. Changes saved successfully."),
    ("Broker Bids", "Delete Bid", "Bid is removed from enquiry. L1/L2 recalculated if necessary. Confirmation shown."),
    ("Broker Bids", "Filter Bids by Enquiry", "When viewing specific enquiry, only relevant bids shown. Filter works correctly."),
    ("Broker Bids", "View Bid Details", "Bid detail shows broker info, rate amount, enquiry details, submission timestamp."),

    # Transport Orders
    ("Transport Orders", "Create Transport Order for Enquiry", "Transport order created linked to enquiry. Broker selected. Route selected. Amount entered. Order appears in list."),
    ("Transport Orders", "Create Multiple Orders for Same Enquiry", "Multiple transport orders can be created for one enquiry. All orders saved. Total cost accumulates."),
    ("Transport Orders", "View All Transport Orders", "All transport orders displayed with enquiry ID, broker, route, amount. Pagination works."),
    ("Transport Orders", "View Transport Order Details", "Order details show enquiry info, broker details, route details, amount, notes, created date."),
    ("Transport Orders", "Select Broker from Dropdown", "Broker dropdown shows all available brokers. Selection works correctly. Selected broker saved."),
    ("Transport Orders", "Select Route from Dropdown", "Route dropdown shows all available routes with locations. Selection works. Route saved to order."),
    ("Transport Orders", "Enter Order Amount", "Amount field accepts numeric input. Validation ensures positive number. Amount saved correctly."),
    ("Transport Orders", "Delete Transport Order", "Order is deleted successfully. Removed from enquiry. Confirmation prompt shown. Quote base amount recalculates."),
    ("Transport Orders", "View Orders from Enquiry Page", "When enquiry expanded, associated transport orders displayed. Broker, route, amount shown."),

    # Quote Management
    ("Quote Management", "Create Calculated Quote - With Margin %", "Quote created with base amount (sum of transport orders), margin percentage, final quotation amount calculated. Formula: base + (base * margin%). Quote status 'pending'."),
    ("Quote Management", "Create Calculated Quote - 10% Margin", "Base amount 10,000, margin 10%, quotation amount = 11,000. Calculation accurate. Quote saved."),
    ("Quote Management", "Create Calculated Quote - 20% Margin", "Base amount 15,000, margin 20%, quotation amount = 18,000. Correct calculation."),
    ("Quote Management", "Create Custom Amount Quote", "Quote created with custom quotation amount (ignores margin calculation). isCustomAmount flag = true. Amount saved as entered."),
    ("Quote Management", "View All Quotes", "All quotes displayed with enquiry reference, lead name, base amount, margin %, quotation amount, status. Pagination works."),
    ("Quote Management", "Filter Quotes by Status - Pending", "Only quotes with status 'pending' displayed. Filter accurate."),
    ("Quote Management", "Filter Quotes by Status - Accepted", "Only accepted quotes shown."),
    ("Quote Management", "Filter Quotes by Status - Rejected", "Only rejected quotes displayed."),
    ("Quote Management", "Filter Quotes by Enquiry ID", "Quotes for specific enquiry shown. Other enquiries excluded."),
    ("Quote Management", "Filter Quotes by Date Range", "Quotes created within date range displayed. Dates outside range excluded."),
    ("Quote Management", "View Quote Details", "Quote detail shows all info: enquiry details, lead info, costing breakdown, base amount, margin, quotation amount, status."),
    ("Quote Management", "Get Base Amount from Transport Orders", "API calculates total of all transport order amounts for an enquiry. Sum is accurate. Displayed as base amount for quote."),
    ("Quote Management", "Edit Quote - Update Margin", "Margin percentage updated. Quotation amount recalculates. Changes saved successfully."),
    ("Quote Management", "Edit Quote - Change to Custom Amount", "Quote type changes to custom. Fixed amount set. isCustomAmount = true."),
    ("Quote Management", "Edit Quote - Change Status to Accepted", "Quote status becomes 'accepted'. Status badge updates. Can proceed to create customer order."),
    ("Quote Management", "Edit Quote - Change Status to Rejected", "Quote status becomes 'rejected'. Status updates. Cannot create customer order from this quote."),
    ("Quote Management", "Delete Quote", "Quote deleted successfully. Removed from list. Confirmation shown. Enquiry remains."),
    ("Quote Management", "View Quotes from Enquiry Page", "When enquiry expanded, all associated quotes shown. Amount, margin, status displayed."),

    # Customer Orders (Final Orders)
    ("Customer Orders", "Create Customer Order from Accepted Quote", "Customer order created from enquiry with accepted quote. Transaction updates: creates customer order, changes enquiry status to 'accepted', updates quote status. All changes atomic."),
    ("Customer Orders", "Create Order - Transaction Success", "Customer order appears in orders list. Enquiry status = 'accepted'. Quote status = 'accepted'. All linked correctly."),
    ("Customer Orders", "Attempt to Create Order from Rejected Quote", "System prevents order creation. Error message shown. No order created."),
    ("Customer Orders", "Attempt to Create Order from Pending Quote", "System prevents or warns. Quote must be accepted first. Validation works."),
    ("Customer Orders", "View All Customer Orders", "All customer orders displayed with enquiry, quote, lead info, status, created date. Pagination works."),
    ("Customer Orders", "View Customer Order Details", "Order detail page shows: enquiry info (from/to, cargo), quote details (amount, margin), lead info (customer name, phone), transport orders, order status, notes."),
    ("Customer Orders", "Filter Orders by Status - Active", "Only active orders displayed. Filter works correctly."),
    ("Customer Orders", "Filter Orders by Status - Completed", "Only completed orders shown."),
    ("Customer Orders", "Filter Orders by Status - Cancelled", "Only cancelled orders displayed."),
    ("Customer Orders", "Update Order Status - Active to Completed", "Order status changes to 'completed'. Status badge updates. Timestamp recorded."),
    ("Customer Orders", "Update Order Status - Active to Cancelled", "Order status changes to 'cancelled'. Status updates correctly."),
    ("Customer Orders", "Add Notes to Customer Order", "Notes field updated. Notes saved. Visible in order details."),
    ("Customer Orders", "Delete Customer Order", "Order deleted. Transaction reverses: enquiry status reverts, quote status reverts. Confirmation shown. All changes atomic."),
    ("Customer Orders", "View Transport Orders from Customer Order", "Customer order detail shows all associated transport orders from enquiry. Broker, route, amounts displayed."),
    ("Customer Orders", "Calculate Total Transport Cost", "Sum of all transport order amounts displayed. Compared with quote amount. Margin visible."),
    ("Customer Orders", "View Customer Order Link from Enquiry", "When enquiry has customer order, link/button to view order shown. Click navigates to order detail."),

    # End-to-End Workflows
    ("End-to-End Workflow", "Complete Flow: Lead to Customer Order", "Create lead → Create enquiry → Create transport rate enquiry → Brokers submit bids → Create transport orders with selected brokers → Generate quote with margin → Accept quote → Create customer order. All steps work seamlessly."),
    ("End-to-End Workflow", "Multiple Quotes per Enquiry Workflow", "Create enquiry → Generate quote 1 with 10% margin (rejected) → Generate quote 2 with 15% margin (rejected) → Generate quote 3 with 8% margin (accepted) → Create customer order. All quotes tracked. Final quote creates order."),
    ("End-to-End Workflow", "Bidding Process Workflow", "Create transport rate enquiry → Status: open → Broker 1 submits bid → Broker 2 submits lower bid (becomes L1) → Broker 3 submits bid → Select best bids → Create transport orders → Close enquiry. Status transitions work."),
    ("End-to-End Workflow", "Order Modification Workflow", "Create customer order → Modify order notes → Update status to active → Change status to completed. All updates successful. Audit trail maintained."),
    ("End-to-End Workflow", "Multi-Source Lead Tracking", "Create lead from IndiaMart → Create enquiry → Create lead from Referral → Create enquiry → Track both separately. Source filtering works. Data isolated correctly."),

    # Filters, Search & Pagination
    ("Filters & Search", "Clear All Filters", "When filters cleared, all records displayed. Filter state resets. Results show full dataset."),
    ("Filters & Search", "Combine Search + Filter", "Search text + status filter work together. Results match both criteria. Accurate filtering."),
    ("Filters & Search", "Pagination with Filters Applied", "Pagination works with active filters. Page count reflects filtered results. Navigation correct."),
    ("Filters & Search", "Search with No Results", "When search has no matches, 'No results found' message shown. Page doesn't break. Clear search returns data."),
    ("Filters & Search", "Date Range Filter - Same Day", "Start and end date same day. Only records from that day shown. Filter accurate."),
    ("Filters & Search", "Date Range Filter - Wide Range", "Large date range (e.g., 1 year). All records in range shown. Performance acceptable."),
    ("Filters & Search", "Pagination - Last Page", "Navigate to last page. Correct number of records shown. Next button disabled."),
    ("Filters & Search", "Pagination - First Page", "First page loads. Previous button disabled. Records display correctly."),
    ("Filters & Search", "Change Page Size", "When page size changed (e.g., 10 to 25), records per page updates. Total pages recalculates."),

    # Data Validation
    ("Data Validation", "Submit Empty Required Field", "Form validation prevents submission. Error message shown. Required field highlighted."),
    ("Data Validation", "Enter Negative Amount in Quote", "Validation prevents negative number. Error shown. Amount must be positive."),
    ("Data Validation", "Enter Non-Numeric Value in Amount", "Validation rejects non-numeric input. Error message displayed. Field shows validation state."),
    ("Data Validation", "Enter Invalid Email Format", "Email validation catches invalid format. Error shown. Format example provided."),
    ("Data Validation", "Enter Margin Over 100%", "System accepts (valid business case). Or validates if threshold set. Calculation works correctly."),
    ("Data Validation", "Enter Zero as Weight", "Validation prevents zero or negative weight. Error shown. Must be positive value."),
    ("Data Validation", "Enter Very Large Number", "System handles large numbers correctly. No overflow errors. Database stores accurately."),
    ("Data Validation", "Special Characters in Text Fields", "Special characters handled correctly. No SQL injection. No XSS vulnerabilities. Data stored safely."),
    ("Data Validation", "Very Long Text in Remarks", "Long text accepted up to field limit. Truncation or scrolling works. No UI breaks."),

    # Security & Authorization
    ("Security & Authorization", "Access Data from Different User", "User A cannot access User B's leads/enquiries/brokers. 403 Forbidden error. Data isolation enforced."),
    ("Security & Authorization", "Attempt API Call Without Token", "API returns 401 Unauthorized. Request rejected. No data exposed."),
    ("Security & Authorization", "Attempt API Call with Expired Token", "Token validation fails. 401 error returned. User redirected to login."),
    ("Security & Authorization", "Attempt API Call with Invalid Token", "Token verification fails. 401 Unauthorized. Request blocked."),
    ("Security & Authorization", "SQL Injection Attempt in Search", "Search input sanitized. SQL injection prevented. No database compromise. Safe query execution."),
    ("Security & Authorization", "XSS Attempt in Text Fields", "Input sanitized/escaped. XSS script not executed. Output rendered safely."),
    ("Security & Authorization", "Password Storage", "Password hashed using bcryptjs. Plain text not stored. Hash verification works on login."),
    ("Security & Authorization", "Password Length Validation", "Passwords < 6 characters rejected. Error shown. Minimum length enforced."),

    # Dashboard & Reporting
    ("Dashboard", "View Dashboard Statistics", "Dashboard shows summary: total leads, enquiries, active orders, pending quotes. Counts accurate."),
    ("Dashboard", "Quick Actions on Dashboard", "Quick action buttons navigate to create lead, create enquiry, view orders. Navigation works."),

    # Edge Cases
    ("Edge Cases", "Create Enquiry Without Transport Orders", "Enquiry created. No transport orders yet. Base amount for quote = 0 or shows warning. Quote calculation handles edge case."),
    ("Edge Cases", "Delete Lead After Creating Order", "System prevents deletion or cascades delete with warnings. Data integrity maintained."),
    ("Edge Cases", "Update Enquiry After Quote Accepted", "Enquiry can be updated. Quote remains unchanged unless recalculated. Warning shown if needed."),
    ("Edge Cases", "Two Users Create Same Lead (Same Phone)", "Both users can create lead with same phone. Data isolated by userId. No conflict."),
    ("Edge Cases", "Browser Refresh During Form Submission", "Form handles refresh gracefully. No duplicate submissions. Data consistency maintained."),
    ("Edge Cases", "Network Error During API Call", "Error message shown to user. Request can be retried. No data corruption."),
    ("Edge Cases", "Concurrent Edits by Same User", "Last write wins or optimistic locking. No data loss. User notified if conflict."),
]

# Add test cases to worksheet
row = 2
current_module = ""
for module, action, expected in test_cases:
    # If module changes, apply module styling
    if module != current_module:
        current_module = module
        cell_a = ws.cell(row=row, column=1, value=module)
        cell_a.fill = module_fill
        cell_a.font = module_font
    else:
        cell_a = ws.cell(row=row, column=1, value="")

    cell_b = ws.cell(row=row, column=2, value=action)
    cell_c = ws.cell(row=row, column=3, value=expected)

    # Apply borders and alignment
    for col in range(1, 4):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    row += 1

# Freeze the header row
ws.freeze_panes = "A2"

# Save the workbook
wb.save("D:\\repos\\sahyogi_transport\\Manual_Test_Plan.xlsx")
print(f"Test plan created successfully with {len(test_cases)} test cases!")
print("File saved as: Manual_Test_Plan.xlsx")
